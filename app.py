import streamlit as st
import pandas as pd
import os
import requests
import base64
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.cell_range import CellRange, MultiCellRange

st.set_page_config(page_title="数据提交系统", page_icon="📊", layout="wide")

# 数据目录和文件配置
DATA_DIR = "data"
TEMPLATES_DIR = "templates"
DATA_FILE = os.path.join(DATA_DIR, "submissions.csv")
BACKUP_CSV_FILE = "data/backup_submissions.csv"
EXPORT_PASSWORD = "907"
GITHUB_REPO = "jinhuang1865/jinhuang1865--"

# 确保目录存在
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(TEMPLATES_DIR, exist_ok=True)

# 初始化数据文件（如果不存在）
if not os.path.exists(DATA_FILE):
    df = pd.DataFrame(columns=["提交时间", "模板名称"])
    df.to_csv(DATA_FILE, index=False, encoding="utf-8-sig")

# 备份到本地CSV
def backup_to_local_csv(df):
    df.to_csv(BACKUP_CSV_FILE, index=False, encoding="utf-8-sig")
    return True

# 备份到GitHub（需要设置环境变量GITHUB_TOKEN）
def backup_to_github():
    try:
        token = os.getenv("GITHUB_TOKEN")
        if not token:
            st.warning("⚠️ 未找到 GITHUB_TOKEN 环境变量，请先设置")
            return False
        # 验证Token有效性
        token_check_url = "https://api.github.com/user"
        check_headers = {"Authorization": f"token {token}"}
        check_r = requests.get(token_check_url, headers=check_headers, timeout=10)
        if check_r.status_code != 200:
            st.error(f"⚠️ GITHUB_TOKEN 无效或权限不足，错误代码：{check_r.status_code}")
            return False
        
        repo = GITHUB_REPO
        path = "data/submissions.csv"
        url = f"https://api.github.com/repos/{repo}/contents/{path}"
        headers = {
            "Authorization": f"token {token}",
            "Accept": "application/vnd.github.json"
        }
        # 读取本地数据文件
        if not os.path.exists(DATA_FILE):
            st.error("⚠️ 本地数据文件不存在，无法备份到GitHub")
            return False
        with open(DATA_FILE, "rb") as f:
            content = base64.b64encode(f.read()).decode()
        
        # 获取文件SHA（如果已存在）
        sha = None
        r = requests.get(url, headers=headers, timeout=10)
        if r.status_code == 200:
            sha = r.json()["sha"]
        elif r.status_code != 404:
            st.error(f"⚠️ 获取GitHub文件失败：{r.text}")
            return False
        
        # 准备提交数据
        data = {
            "message": f"Auto backup {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "content": content,
            "branch": "main"
        }
        if sha:
            data["sha"] = sha
        
        # 提交到GitHub
        r = requests.put(url, headers=headers, json=data, timeout=15)
        if r.status_code in [200, 201]:
            st.success("✅ GitHub备份成功！")
            return True
        else:
            st.error(f"❌ GitHub API提交失败：{r.status_code}，错误：{r.text}")
            return False
    except requests.exceptions.Timeout:
        st.error("❌ GitHub连接超时，请检查网络连接")
        return False
    except Exception as e:
        st.error(f"❌ GitHub备份错误：{str(e)}")
        return False

# 获取模板文件列表（排除临时文件）
def get_template_files():
    templates = []
    for f in os.listdir(TEMPLATES_DIR):
        if f.endswith('.xlsx') and not f.startswith('~'):
            templates.append(f)
    return sorted(templates)

# 获取模板列名（排除时间和模板名列）
def get_template_columns(template_name):
    template_path = os.path.join(TEMPLATES_DIR, template_name)
    try:
        df = pd.read_excel(template_path)
        return [col for col in df.columns if col not in ["提交时间", "模板名称"]]
    except Exception as e:
        st.error(f"读取模板失败：{e}")
        return []

# 从模板中提取下拉选项（支持CellRange/MultiCellRange等复杂情况）
def get_dropdown_options_from_template(template_path):
    options_dict = {}
    try:
        wb = load_workbook(template_path, data_only=True)
        ws = wb.active
        if not ws.data_validations:
            return options_dict  # 无数据验证
        
        for dv in ws.data_validations.dataValidation:
            col = None
            # 处理单元格范围：支持CellRange/MultiCellRange等
            if isinstance(dv.cells, CellRange):
                # 单个CellRange：取最小列
                col = dv.cells.min_col
            elif isinstance(dv.cells, MultiCellRange):
                # MultiCellRange：取第一个范围的列
                if dv.cells.ranges:
                    col = dv.cells.ranges[0].min_col
            else:
                # 其他情况跳过
                continue
            
            if not col:
                continue  # 无法确定列
            
            # 获取列标题（假设第一行是标题）
            title_cell = ws.cell(row=1, column=col)
            col_name = title_cell.value
            if not col_name or col_name.strip() == "":
                continue  # 无标题跳过
            
            # 提取验证选项
            formula = dv.formula1
            if not formula:
                continue  # 无公式跳过
            options = []
            
            # 处理选项字符串（如："选项1,选项2,选项3"）
            if formula.startswith('"') and formula.endswith('"'):
                options_str = formula[1:-1]
                options = [opt.strip() for opt in options_str.split(',') if opt.strip()]
            # 处理引用范围（如：'Sheet1'!$A$1:$A$5）
            elif '!' in formula:
                try:
                    # 分离工作表名和范围
                    sheet_part, range_part = formula.split('!', 1)
                    sheet_name = sheet_part.strip("'")
                    # 解析范围
                    min_c, min_r, max_c, max_r = range_boundaries(range_part)
                    # 获取引用工作表
                    ref_ws = wb[sheet_name] if sheet_name in wb.sheetnames else ws
                    # 遍历范围获取选项值
                    for row in ref_ws.iter_rows(min_col=min_c, max_col=max_c, min_row=min_r, max_row=max_r):
                        for cell in row:
                            val = cell.value
                            if val and str(val).strip() != "":
                                options.append(str(val).strip())
                    # 去重
                    options = list(dict.fromkeys(options))
                except Exception as e:
                    # 解析失败跳过
                    continue
            
            # 保存选项到字典
            if options:
                options_dict[col_name] = options
    except Exception as e:
        # 错误处理
        pass
    return options_dict

# 清理空数据行/空值
def clean_empty_data(df):
    # 删除全空行
    df = df.dropna(how='all')
    # 删除全空白字符的行
    df = df[df.apply(lambda row: row.astype(str).str.strip().any(), axis=1)]
    # 重置索引
    df = df.reset_index(drop=True)
    return df

# 主界面设置
st.title("📊 数据提交系统")
st.markdown("---")

# 功能选项卡（模板管理/数据提交/数据查看/数据导出）
tab1, tab2, tab3, tab4 = st.tabs([
    "📁 模板管理",
    "📥 数据提交",
    "📋 数据查看",
    "📈 数据导出"
])

# --- 模板管理选项卡 ---
with tab1:
    st.header("📁 模板管理")
    admin_password = st.text_input("🔐 管理员密码", type="password", key="tab1_password")
    if admin_password != EXPORT_PASSWORD:
        st.warning("🔒 请输入正确的管理员密码")
    else:
        st.success("✅ 验证成功")
        new_template_file = st.file_uploader("📄 上传Excel模板文件（支持.xlsx格式）", type=["xlsx"], key="upload_template")
        if new_template_file:
            template_name = st.text_input("📁 模板名称（不含扩展名）", key="template_name", placeholder="例如：2024年度数据收集")
            if st.button("💾 保存模板", key="save_template"):
                if template_name:
                    original_ext = os.path.splitext(new_template_file.name)[1]
                    template_path = os.path.join(TEMPLATES_DIR, f"{template_name}{original_ext}")
                    with open(template_path, "wb") as f:
                        f.write(new_template_file.getbuffer())
                    st.success(f"✅ 模板 '{template_name}' 保存成功！")
                    st.rerun()
                else:
                    st.error("❌ 请输入模板名称")
        st.markdown("---")
        st.subheader("📊 现有的模板")
        templates = get_template_files()
        if templates:
            for t in templates:
                col1, col2 = st.columns([4, 1])
                col1.write(f"📄 {t}")
                if col2.button("🗑️ 删除", key=f"delete_{t}"):
                    os.remove(os.path.join(TEMPLATES_DIR, t))
                    st.success(f"✅ 删除 {t}")
                    st.rerun()
        else:
            st.info("暂无模板")

# --- 数据提交选项卡 ---
with tab2:
    st.header("📥 数据提交")
    templates = get_template_files()
    if not templates:
        st.warning("⚠️ 暂无可用模板")
    else:
        selected_template = st.selectbox("📊 选择模板", templates, key="download_template")
        columns = get_template_columns(selected_template)
        if columns:
            st.info(f"📋 模板包含列：{', '.join(columns)}")
        template_path = os.path.join(TEMPLATES_DIR, selected_template)
        with open(template_path, "rb") as f:
            st.download_button(
                label="📥 下载模板",
                data=f,
                file_name=selected_template,
                key="download_template_btn"
            )

# --- 数据查看选项卡 ---
with tab3:
    st.header("📋 数据查看")
    templates = get_template_files()
    if not templates:
        st.warning("⚠️ 暂无可用模板")
    else:
        selected_template = st.selectbox("📊 选择模板", templates, key="upload_template_select")
        template_path = os.path.join(TEMPLATES_DIR, selected_template)
        template_cols = get_template_columns(selected_template)
        dropdown_options = get_dropdown_options_from_template(template_path)
        
        if dropdown_options:
            st.info(f"📝 模板包含下拉选项：{', '.join(dropdown_options.keys())}")
        
        uploaded_file = st.file_uploader("📄 上传Excel数据文件（支持.xlsx/.xls格式）", type=["xlsx", "xls"], key="upload_excel")
        # 使用session_state管理文件上传状态
        if "uploaded_file_key" not in st.session_state:
            st.session_state.uploaded_file_key = None
        
        if uploaded_file:
            # 生成文件唯一标识（文件名+大小）
            current_file_key = f"{uploaded_file.name}_{uploaded_file.size}"
            if st.session_state.uploaded_file_key != current_file_key:
                try:
                    df_upload = pd.read_excel(uploaded_file)
                    # 清理空数据行/空值
                    df_upload = clean_empty_data(df_upload)
                    if len(df_upload) == 0:
                        st.warning("⚠️ 上传的文件没有有效数据（空行/空值过多），请重新上传")
                        st.session_state.uploaded_file_key = current_file_key
                        st.stop()
                    
                    # 1. 检查必填列是否存在
                    missing_cols = [col for col in template_cols if col not in df_upload.columns]
                    if missing_cols:
                        st.error(f"❌ 文件缺少必填列：{', '.join(missing_cols)}，请重新上传")
                        st.session_state.uploaded_file_key = current_file_key
                        st.stop()
                    
                    # 2. 检查空值列（所有行都为空）
                    empty_cols = []
                    for col in template_cols:
                        if col in df_upload.columns:
                            # 检查是否为NaN或空字符串
                            if df_upload[col].isna().any() or (df_upload[col].astype(str).str.strip() == '').any():
                                empty_cols.append(col)
                    if empty_cols:
                        st.error(f"❌ 文件存在空值列，请重新上传：{', '.join(empty_cols)}")
                        st.session_state.uploaded_file_key = current_file_key
                        st.stop()
                    
                    # 3. 验证下拉选项合规性
                    invalid_options = []
                    if dropdown_options:
                        for col, options in dropdown_options.items():
                            if col in df_upload.columns:
                                values = df_upload[col].dropna().astype(str).str.strip()
                                values = values[~values.str.contains("请选择", na=False)]
                                invalid = [v for v in values if v not in options]
                                if invalid:
                                    invalid_options.append(f"{col}：{','.join(invalid[:3])}等{len(invalid)}个无效值")
                    if invalid_options:
                        st.error(f"❌ 数据存在无效选项：{'; '.join(invalid_options)}")
                        st.info("💡 请参考模板中的下拉选项重新填写")
                        st.session_state.uploaded_file_key = current_file_key
                        st.stop()
                    
                    # 添加模板和时间标识
                    df_upload["模板名称"] = selected_template
                    df_upload["提交时间"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    
                    # 合并到主数据文件
                    df_existing = pd.read_csv(DATA_FILE, encoding="utf-8-sig")
                    df_combined = pd.concat([df_existing, df_upload], ignore_index=True)
                    # 去重处理（基于"身份证"/"学号"等唯一标识）
                    if "身份证" in df_combined.columns:
                        df_combined = df_combined.drop_duplicates(subset=["身份证"], keep="last")
                    elif "学号" in df_combined.columns:
                        df_combined = df_combined.drop_duplicates(subset=["学号"], keep="last")
                    
                    # 保存到主文件
                    df_combined.to_csv(DATA_FILE, index=False, encoding="utf-8-sig")
                    # 备份到本地CSV
                    backup_to_local_csv(df_combined)
                    # 备份到GitHub
                    with st.spinner("🔄 正在备份到GitHub..."):
                        backup_to_github()
                    
                    st.success(f"✅ 数据提交成功！本次上传 {len(df_upload)} 条有效记录")
                    st.dataframe(df_upload.head(10), use_container_width=True)  # 显示前10条数据
                    
                    # 更新session_state状态
                    st.session_state.uploaded_file_key = current_file_key
                    
                except Exception as e:
                    st.error(f"❌ 文件处理失败：{str(e)}")
            else:
                # 已处理过相同文件
                st.info("ℹ️ 已上传过相同文件，如需重新上传请更换文件")
        else:
            st.info("📤 请上传数据文件开始提交")

# --- 数据导出选项卡 ---
with tab4:
    st.header("📈 数据导出")
    password = st.text_input("🔐 导出密码", type="password", key="tab4_password")
    if password == EXPORT_PASSWORD:
        df_all = pd.read_csv(DATA_FILE, encoding="utf-8-sig")
        if len(df_all) == 0:
            st.warning("⚠️ 暂无提交数据")
        else:
            # ── 按模板筛选数据 ────────────────────────────────────
            template_filter = "全部数据"
            if "模板名称" in df_all.columns:
                template_filter = st.selectbox(
                    "📊 按模板筛选",
                    ["全部数据"] + list(df_all["模板名称"].unique()),
                    key="template_filter"
                )
                if template_filter != "全部数据":
                    df_all = df_all[df_all["模板名称"] == template_filter].reset_index(drop=True)

            # ── 修复：按所选模板严格过滤表头列 ───────────────────
            # 当选定了某个模板时，只保留该模板对应的列 + 提交时间 + 模板名称
            if template_filter != "全部数据":
                template_cols = get_template_columns(template_filter)
                # 保留模板列 + 系统列（提交时间、模板名称），去掉其他模板带来的多余列
                keep_cols = ["提交时间", "模板名称"] + template_cols
                keep_cols = [c for c in keep_cols if c in df_all.columns]
                df_all = df_all[keep_cols]

            # ── 数据统计指标 ──────────────────────────────────────
            dept_col = None
            if "二级部门" in df_all.columns:
                dept_col = "二级部门"
            else:
                dept_candidates = [c for c in df_all.columns if "部门" in str(c)]
                if dept_candidates:
                    dept_col = dept_candidates[0]

            col1, col2, col3, col4 = st.columns(4)
            col1.metric("总记录数", len(df_all))
            col2.metric("模板数", df_all["模板名称"].nunique() if "模板名称" in df_all.columns else 0)
            if dept_col:
                col3.metric(f"{dept_col}数量", df_all[dept_col].nunique())
            else:
                col3.metric("部门数量", "—")
            col4.metric("最新提交", df_all["提交时间"].max() if "提交时间" in df_all.columns else "N/A")

            # ── 数据表格 ──────────────────────────────────────────
            st.markdown("---")
            st.subheader("📋 数据列表")

            # 生成带序号的显示表格（避免重复插入序号列）
            df_display = df_all.reset_index(drop=True)
            if "序号" in df_display.columns:
                df_display = df_display.drop(columns=["序号"])
            df_display.insert(0, "序号", range(1, len(df_display) + 1))

            st.dataframe(df_display, use_container_width=True, height=400)

            # ── 删除记录区域 ──────────────────────────────────────
            st.markdown("#### 🗑️ 删除记录")

            del_tab1, del_tab2 = st.tabs(["按序号删除（单条/多条）", "按范围批量删除"])

            # ---- 单条/多条删除 ----
            with del_tab1:
                st.caption("输入要删除的序号，多个用英文逗号分隔，例如：1,3,5")
                delete_input = st.text_input("要删除的序号", placeholder="例如：1,3,5", key="delete_input")
                if st.button("🗑️ 确认删除", key="confirm_delete_single"):
                    if delete_input.strip():
                        try:
                            input_nums = [int(x.strip()) for x in delete_input.split(",") if x.strip()]
                            indices_to_delete = [n - 1 for n in input_nums if 1 <= n <= len(df_display)]
                            if not indices_to_delete:
                                st.error("❌ 未找到有效序号，请检查输入")
                            else:
                                df_origin = pd.read_csv(DATA_FILE, encoding="utf-8-sig")
                                filtered_real_indices = df_all.index.tolist()
                                real_indices_to_delete = [filtered_real_indices[i] for i in indices_to_delete]
                                df_origin = df_origin.drop(index=real_indices_to_delete).reset_index(drop=True)
                                df_origin.to_csv(DATA_FILE, index=False, encoding="utf-8-sig")
                                backup_to_local_csv(df_origin)
                                with st.spinner("🔄 正在同步到GitHub..."):
                                    backup_to_github()
                                st.success(f"✅ 已成功删除 {len(real_indices_to_delete)} 条记录")
                                st.rerun()
                        except ValueError:
                            st.error("❌ 序号格式有误，请输入数字，多个用英文逗号分隔")
                    else:
                        st.warning("⚠️ 请先输入要删除的序号")

            # ---- 范围批量删除 ----
            with del_tab2:
                st.caption("输入起止序号范围，例如：1-10，将删除序号1至10的全部数据")
                range_input = st.text_input("序号范围", placeholder="例如：1-10", key="delete_range_input")

                # 初始化确认状态
                if "batch_delete_confirm" not in st.session_state:
                    st.session_state.batch_delete_confirm = False
                if "batch_delete_range" not in st.session_state:
                    st.session_state.batch_delete_range = ""

                col_r1, col_r2 = st.columns([1, 5])
                with col_r1:
                    if st.button("🗑️ 批量删除", key="batch_delete_btn"):
                        if range_input.strip():
                            try:
                                parts = range_input.strip().split("-")
                                if len(parts) != 2:
                                    raise ValueError
                                start_num = int(parts[0].strip())
                                end_num = int(parts[1].strip())
                                if start_num < 1 or end_num < start_num or end_num > len(df_display):
                                    st.error(f"❌ 序号范围无效，请输入 1 到 {len(df_display)} 之间的范围")
                                else:
                                    # 触发确认弹窗
                                    st.session_state.batch_delete_confirm = True
                                    st.session_state.batch_delete_range = f"{start_num}-{end_num}"
                                    st.rerun()
                            except ValueError:
                                st.error("❌ 格式有误，请输入如：1-10 的范围")
                        else:
                            st.warning("⚠️ 请先输入序号范围")

                # ── 确认弹窗（用 st.dialog 模拟） ─────────────────
                if st.session_state.get("batch_delete_confirm", False):
                    rng = st.session_state.get("batch_delete_range", "")
                    if rng:
                        parts = rng.split("-")
                        s_num, e_num = int(parts[0]), int(parts[1])
                        count = e_num - s_num + 1
                        st.warning(
                            f"⚠️ **即将删除序号 {s_num}-{e_num} 的 {count} 条数据，此操作不可恢复，是否确认？**"
                        )
                        c_yes, c_no, _ = st.columns([1, 1, 4])
                        with c_yes:
                            if st.button("✅ 确认删除", key="batch_confirm_yes"):
                                try:
                                    indices_to_delete = list(range(s_num - 1, e_num))
                                    df_origin = pd.read_csv(DATA_FILE, encoding="utf-8-sig")
                                    filtered_real_indices = df_all.index.tolist()
                                    real_indices_to_delete = [filtered_real_indices[i] for i in indices_to_delete]
                                    df_origin = df_origin.drop(index=real_indices_to_delete).reset_index(drop=True)
                                    df_origin.to_csv(DATA_FILE, index=False, encoding="utf-8-sig")
                                    backup_to_local_csv(df_origin)
                                    with st.spinner("🔄 正在同步到GitHub..."):
                                        backup_to_github()
                                    st.session_state.batch_delete_confirm = False
                                    st.session_state.batch_delete_range = ""
                                    st.success(f"✅ 已成功批量删除序号 {s_num}-{e_num} 共 {count} 条记录")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"❌ 删除失败：{str(e)}")
                        with c_no:
                            if st.button("❌ 取消", key="batch_confirm_no"):
                                st.session_state.batch_delete_confirm = False
                                st.session_state.batch_delete_range = ""
                                st.rerun()

            st.markdown("---")

            # ── 导出Excel（严格按当前表头导出）──────────────────────
            export_filename = f"数据导出_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df_all.to_excel(writer, index=False, sheet_name="数据")
            excel_data = output.getvalue()

            st.download_button(
                label="📄 导出Excel",
                data=excel_data,
                file_name=export_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="export_excel"
            )
    else:
        st.warning("🔒 请输入正确的导出密码")
